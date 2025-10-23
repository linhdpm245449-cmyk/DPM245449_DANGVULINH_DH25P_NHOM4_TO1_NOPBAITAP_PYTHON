
def employee_excel_manager():
    try:
        from openpyxl import Workbook, load_workbook
    except:
        path = "employees.csv"
        while True:
            print("\n1-them nhan vien 2-doc danh sach 3-sap xep theo tuoi 4-thoat")
            c = input("chon: ").strip()
            if c == '1':
                masv = input("Ma NV: ").strip()
                name = input("Ten NV: ").strip()
                age = int(input("Tuoi: "))
                with open(path, 'a', encoding='utf-8', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow([masv, name, age])
                print("Da them")
            elif c == '2':
                if not os.path.exists(path):
                    print("Chua co file")
                    continue
                with open(path, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    for row in reader:
                        print(row)
            elif c == '3':
                if not os.path.exists(path):
                    print("Chua co file")
                    continue
                rows = []
                with open(path, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    for row in reader:
                        if len(row) >= 3:
                            rows.append([row[0], row[1], int(row[2])])
                rows.sort(key=lambda x: x[2])
                for r in rows:
                    print(r)
            elif c == '4':
                break
            else:
                print("Lua chon khong hop le")
        return

    path = "employees.xlsx"
    def ensure_workbook():
        if not os.path.exists(path):
            wb = Workbook()
            ws = wb.active
            ws.append(["Ma", "Ten", "Tuoi"])
            wb.save(path)
    ensure_workbook()
    while True:
        print("\n1-them nhan vien 2-doc danh sach 3-sap xep theo tuoi 4-thoat")
        c = input("chon: ").strip()
        if c == '1':
            masv = input("Ma NV: ").strip()
            name = input("Ten NV: ").strip()
            age = int(input("Tuoi: "))
            wb = load_workbook(path)
            ws = wb.active
            ws.append([masv, name, age])
            wb.save(path)
            print("Da them")
        elif c == '2':
            wb = load_workbook(path)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                print(row)
        elif c == '3':
            wb = load_workbook(path)
            ws = wb.active
            rows = []
            first = True
            for row in ws.iter_rows(values_only=True):
                if first:
                    first = False
                    continue
                rows.append([row[0], row[1], row[2]])
            rows.sort(key=lambda x: x[2])
            wb2 = Workbook()
            ws2 = wb2.active
            ws2.append(["Ma", "Ten", "Tuoi"])
            for r in rows:
                ws2.append(r)
            wb2.save(path)
            print("Da sap xep va luu")
        elif c == '4':
            break
        else:
            print("Lua chon khong hop le")
